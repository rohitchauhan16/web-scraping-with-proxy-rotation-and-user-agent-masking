import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import openpyxl

# Women Shoes
# Define the base URL of the website
base_url = 'https://singapore.coach.com/women/shoes.html'

all_links = []
page_num = 1

while True:
    # Construct the URL for the current page
    url = f"{base_url}?p={page_num}"
    
    # Send a GET request to the website
    response = requests.get(url)
    
    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content of the webpage
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Find the specific tab content
        tab_contents = soup.findAll(class_='product details product-item-details')
        
        if not tab_contents:
            break
        
        for tab_content in tab_contents:
            # Find all 'a' tags within the tab content
            links = tab_content.find_all('a', href=True)
        
            # Extract and print the 'href' attributes
            for link in links:
                all_links.append(link['href'])
                print(link['href'])
        
        # Check for the presence of a next page link (example assumes a 'next' button or page number link)
        next_page = soup.find('a', {'class': 'next'})  # Adjust the class or criteria as needed
        if not next_page:
            break
        
        # Increment page number for the next iteration
        page_num += 1
    else:
        break

print("Total links found:", len(all_links))


al=pd.DataFrame(all_links, columns=['links'])
al

# Initialize an empty DataFrame
ws = pd.DataFrame(columns=[
    'Link', 'Name', 'Image', 'Category', 'Subcategory', 'Color', 'Current Price', 'Original Price', 'Description', 'Details', 
    'Tag', 'Status', 'Remarks', 'Timestamp'
])

# Loop through each URL and scrape data
for url in al.links.values:
    response = requests.get(url)
    
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')

        # Extract the product name
        product_link = url

        # Extract the product name
        product_name_tag = soup.find('span', class_='base')
        product_name = product_name_tag.text.strip() if product_name_tag else 'No Product Name Available'
        
        # Extract the image URL
        img_tag = soup.find('img', class_='gallery-placeholder__image') 
        img_url = img_tag['src'] if img_tag and 'src' in img_tag.attrs else 'No Image Available'
        
        # Extract the category
        # category_tag = soup.find('span', class_='category')
        # category = category_tag.text.strip() if category_tag else 'No Category Available'
        category= 'Women - Shoes'

        # Extract the subcategory
        # subcategory_tag = soup.find('span', class_='subcategory')
        # subcategory = subcategory_tag.text.strip() if subcategory_tag else 'No Subcategory Available'
        subcategory= ''

        # Extract the color
        product_color_tag = soup.find('div', class_='value', itemprop='coach_style_color') 
        product_color = product_color_tag.text.strip() if product_color_tag else 'No Color Available'
        
        # Extract the current price
        current_price_tag = soup.find('span', class_='price-wrapper', attrs={'data-price-type': "finalPrice"})
        current_price = current_price_tag.text.strip() if current_price_tag else 'No Current Price Available'
        
        # Extract the original price
        original_price_tag = soup.find('span', class_='price-wrapper', attrs={'data-price-type': "oldPrice"}) 
        original_price = original_price_tag.text.strip() if original_price_tag else 'No Original Price Available'
        
        # Extract the description
        description_tag = soup.find('div', class_='product attribute overview')
        description = description_tag.text.strip() if description_tag else 'No Description Available'
        
        # Extract the details
        details_tag = soup.find('div', class_='product attribute description')
        details = details_tag.text.strip() if details_tag else 'No Details Available'
        
        # Extract the tag (internal links within the product page)
        tag_tag = soup.find('a', class_='internal-link')
        tag = tag_tag['href'] if tag_tag and 'href' in tag_tag.attrs else 'No Tag Available'
        
        # Extract the status (OOS/Pre-order, etc.)
        status_tag = soup.find('div', class_='status')
        status = status_tag.text.strip() if status_tag else 'No Status Available'
        
        # Extract the remarks (Stock arrival period)
        remarks_tag = soup.find('div', class_='product-info-stock-sku')
        remarks = remarks_tag.text.strip() if remarks_tag else 'No Remarks Available'
        
        # Get the current timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create a DataFrame for the current product
        data = {
            'Link': [product_link],
            'Name': [product_name],
            'Image': [img_url],
            'Category': [category],
            'Subcategory': [subcategory],
            'Color': [product_color],
            'Current Price': [current_price],
            'Original Price': [original_price],
            'Description': [description],
            'Details': [details],
            'Tag': [tag],
            'Status': [status],
            'Remarks': [remarks],
            'Timestamp': [timestamp]
        }
        temp_ws = pd.DataFrame(data)
        
        # Concatenate the current product DataFrame to the main DataFrame
        ws = pd.concat([ws, temp_ws], ignore_index=True)
    else:
        print(f'Failed to retrieve the webpage: {url}')

# Display the DataFrame
# print(ws)
# ws

ws['Remarks']=ws['Remarks'].transform(lambda x: ' '.join(x.split()).strip() if x else x)
ws['Details']=ws['Details'].transform(lambda x: x.replace('\n', ',') if x else x)
ws 


# Women Bags
# Define the base URL of the website
base_url = 'https://singapore.coach.com/women/bags.html'

all_links = []
page_num = 1

while True:
    # Construct the URL for the current page
    url = f"{base_url}?p={page_num}"
    
    # Send a GET request to the website
    response = requests.get(url)
    
    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content of the webpage
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Find the specific tab content
        tab_contents = soup.findAll(class_='product details product-item-details')
        
        if not tab_contents:
            break
        
        for tab_content in tab_contents:
            # Find all 'a' tags within the tab content
            links = tab_content.find_all('a', href=True)
        
            # Extract and print the 'href' attributes
            for link in links:
                all_links.append(link['href'])
                print(link['href'])
        
        # Check for the presence of a next page link (example assumes a 'next' button or page number link)
        next_page = soup.find('a', {'class': 'next'})  # Adjust the class or criteria as needed
        if not next_page:
            break
        
        # Increment page number for the next iteration
        page_num += 1
    else:
        break

print("Total links found:", len(all_links))


al=pd.DataFrame(all_links, columns=['links'])
al

# Initialize an empty DataFrame
wb = pd.DataFrame(columns=[
    'Link', 'Name', 'Image', 'Category', 'Subcategory', 'Color', 'Current Price', 'Original Price', 'Description', 'Details', 
    'Tag', 'Status', 'Remarks', 'Timestamp'
])

# Loop through each URL and scrape data
for url in al.links.values:
    response = requests.get(url)
    
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')

        # Extract the product name
        product_link = url

        # Extract the product name
        product_name_tag = soup.find('span', class_='base')
        product_name = product_name_tag.text.strip() if product_name_tag else 'No Product Name Available'
        
        # Extract the image URL
        img_tag = soup.find('img', class_='gallery-placeholder__image') 
        img_url = img_tag['src'] if img_tag and 'src' in img_tag.attrs else 'No Image Available'
        
        # Extract the category
        # category_tag = soup.find('span', class_='category')
        # category = category_tag.text.strip() if category_tag else 'No Category Available'
        category= 'Women - Bags'

        # Extract the subcategory
        # subcategory_tag = soup.find('span', class_='subcategory')
        # subcategory = subcategory_tag.text.strip() if subcategory_tag else 'No Subcategory Available'
        subcategory= ''

        # Extract the color
        product_color_tag = soup.find('div', class_='value', itemprop='coach_style_color') 
        product_color = product_color_tag.text.strip() if product_color_tag else 'No Color Available'
        
        # Extract the current price
        current_price_tag = soup.find('span', class_='price-wrapper', attrs={'data-price-type': "finalPrice"})
        current_price = current_price_tag.text.strip() if current_price_tag else 'No Current Price Available'
        
        # Extract the original price
        original_price_tag = soup.find('span', class_='price-wrapper', attrs={'data-price-type': "oldPrice"}) 
        original_price = original_price_tag.text.strip() if original_price_tag else 'No Original Price Available'
        
        # Extract the description
        description_tag = soup.find('div', class_='product attribute overview')
        description = description_tag.text.strip() if description_tag else 'No Description Available'
        
        # Extract the details
        details_tag = soup.find('div', class_='product attribute description')
        details = details_tag.text.strip() if details_tag else 'No Details Available'
        
        # Extract the tag (internal links within the product page)
        tag_tag = soup.find('a', class_='internal-link')
        tag = tag_tag['href'] if tag_tag and 'href' in tag_tag.attrs else 'No Tag Available'
        
        # Extract the status (OOS/Pre-order, etc.)
        status_tag = soup.find('div', class_='status')
        status = status_tag.text.strip() if status_tag else 'No Status Available'
        
        # Extract the remarks (Stock arrival period)
        remarks_tag = soup.find('div', class_='product-info-stock-sku')
        remarks = remarks_tag.text.strip() if remarks_tag else 'No Remarks Available'
        
        # Get the current timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create a DataFrame for the current product
        data = {
            'Link': [product_link],
            'Name': [product_name],
            'Image': [img_url],
            'Category': [category],
            'Subcategory': [subcategory],
            'Color': [product_color],
            'Current Price': [current_price],
            'Original Price': [original_price],
            'Description': [description],
            'Details': [details],
            'Tag': [tag],
            'Status': [status],
            'Remarks': [remarks],
            'Timestamp': [timestamp]
        }
        temp_wb = pd.DataFrame(data)
        
        # Concatenate the current product DataFrame to the main DataFrame
        wb = pd.concat([wb, temp_wb], ignore_index=True)
    else:
        print(f'Failed to retrieve the webpage: {url}')

# Display the DataFrame
# print(wb)
# wb

wb['Remarks']=wb['Remarks'].transform(lambda x: ' '.join(x.split()).strip() if x else x)
wb['Details']=wb['Details'].transform(lambda x: x.replace('\n', ',') if x else x)
wb 



# Men Shoes
# Define the base URL of the website
base_url = 'https://singapore.coach.com/men/shoes.html'

all_links = []
page_num = 1

while True:
    # Construct the URL for the current page
    url = f"{base_url}?p={page_num}"
    
    # Send a GET request to the website
    response = requests.get(url)
    
    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content of the webpage
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Find the specific tab content
        tab_contents = soup.findAll(class_='product details product-item-details')
        
        if not tab_contents:
            break
        
        for tab_content in tab_contents:
            # Find all 'a' tags within the tab content
            links = tab_content.find_all('a', href=True)
        
            # Extract and print the 'href' attributes
            for link in links:
                all_links.append(link['href'])
                print(link['href'])
        
        # Check for the presence of a next page link (example assumes a 'next' button or page number link)
        next_page = soup.find('a', {'class': 'next'})  # Adjust the class or criteria as needed
        if not next_page:
            break
        
        # Increment page number for the next iteration
        page_num += 1
    else:
        break

print("Total links found:", len(all_links))


al=pd.DataFrame(all_links, columns=['links'])
al

# Initialize an empty DataFrame
ms = pd.DataFrame(columns=[
    'Link', 'Name', 'Image', 'Category', 'Subcategory', 'Color', 'Current Price', 'Original Price', 'Description', 'Details', 
    'Tag', 'Status', 'Remarks', 'Timestamp'
])

# Loop through each URL and scrape data
for url in al.links.values:
    response = requests.get(url)
    
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')

        # Extract the product name
        product_link = url

        # Extract the product name
        product_name_tag = soup.find('span', class_='base')
        product_name = product_name_tag.text.strip() if product_name_tag else 'No Product Name Available'
        
        # Extract the image URL
        img_tag = soup.find('img', class_='gallery-placeholder__image') 
        img_url = img_tag['src'] if img_tag and 'src' in img_tag.attrs else 'No Image Available'
        
        # Extract the category
        # category_tag = soup.find('span', class_='category')
        # category = category_tag.text.strip() if category_tag else 'No Category Available'
        category= 'Men - Shoes'

        # Extract the subcategory
        # subcategory_tag = soup.find('span', class_='subcategory')
        # subcategory = subcategory_tag.text.strip() if subcategory_tag else 'No Subcategory Available'
        subcategory= ''

        # Extract the color
        product_color_tag = soup.find('div', class_='value', itemprop='coach_style_color') 
        product_color = product_color_tag.text.strip() if product_color_tag else 'No Color Available'
        
        # Extract the current price
        current_price_tag = soup.find('span', class_='price-wrapper', attrs={'data-price-type': "finalPrice"})
        current_price = current_price_tag.text.strip() if current_price_tag else 'No Current Price Available'
        
        # Extract the original price
        original_price_tag = soup.find('span', class_='price-wrapper', attrs={'data-price-type': "oldPrice"}) 
        original_price = original_price_tag.text.strip() if original_price_tag else 'No Original Price Available'
        
        # Extract the description
        description_tag = soup.find('div', class_='product attribute overview')
        description = description_tag.text.strip() if description_tag else 'No Description Available'
        
        # Extract the details
        details_tag = soup.find('div', class_='product attribute description')
        details = details_tag.text.strip() if details_tag else 'No Details Available'
        
        # Extract the tag (internal links within the product page)
        tag_tag = soup.find('a', class_='internal-link')
        tag = tag_tag['href'] if tag_tag and 'href' in tag_tag.attrs else 'No Tag Available'
        
        # Extract the status (OOS/Pre-order, etc.)
        status_tag = soup.find('div', class_='status')
        status = status_tag.text.strip() if status_tag else 'No Status Available'
        
        # Extract the remarks (Stock arrival period)
        remarks_tag = soup.find('div', class_='product-info-stock-sku')
        remarks = remarks_tag.text.strip() if remarks_tag else 'No Remarks Available'
        
        # Get the current timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create a DataFrame for the current product
        data = {
            'Link': [product_link],
            'Name': [product_name],
            'Image': [img_url],
            'Category': [category],
            'Subcategory': [subcategory],
            'Color': [product_color],
            'Current Price': [current_price],
            'Original Price': [original_price],
            'Description': [description],
            'Details': [details],
            'Tag': [tag],
            'Status': [status],
            'Remarks': [remarks],
            'Timestamp': [timestamp]
        }
        temp_ms = pd.DataFrame(data)
        
        # Concatenate the current product DataFrame to the main DataFrame
        ms = pd.concat([ms, temp_ms], ignore_index=True)
    else:
        print(f'Failed to retrieve the webpage: {url}')

# Display the DataFrame
# print(ms)
# ms

ms['Remarks']=ms['Remarks'].transform(lambda x: ' '.join(x.split()).strip() if x else x)
ms['Details']=ms['Details'].transform(lambda x: x.replace('\n', ',') if x else x)
ms 



# Men Bags
# Define the base URL of the website
base_url = 'https://singapore.coach.com/men/bag.html'

all_links = []
page_num = 1

while True:
    # Construct the URL for the current page
    url = f"{base_url}?p={page_num}"
    
    # Send a GET request to the website
    response = requests.get(url)
    
    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content of the webpage
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Find the specific tab content
        tab_contents = soup.findAll(class_='product details product-item-details')
        
        if not tab_contents:
            break
        
        for tab_content in tab_contents:
            # Find all 'a' tags within the tab content
            links = tab_content.find_all('a', href=True)
        
            # Extract and print the 'href' attributes
            for link in links:
                all_links.append(link['href'])
                print(link['href'])
        
        # Check for the presence of a next page link (example assumes a 'next' button or page number link)
        next_page = soup.find('a', {'class': 'next'})  # Adjust the class or criteria as needed
        if not next_page:
            break
        
        # Increment page number for the next iteration
        page_num += 1
    else:
        break

print("Total links found:", len(all_links))


al=pd.DataFrame(all_links, columns=['links'])
al

# Initialize an empty DataFrame
mb = pd.DataFrame(columns=[
    'Link', 'Name', 'Image', 'Category', 'Subcategory', 'Color', 'Current Price', 'Original Price', 'Description', 'Details', 
    'Tag', 'Status', 'Remarks', 'Timestamp'
])

# Loop through each URL and scrape data
for url in al.links.values:
    response = requests.get(url)
    
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')

        # Extract the product name
        product_link = url

        # Extract the product name
        product_name_tag = soup.find('span', class_='base')
        product_name = product_name_tag.text.strip() if product_name_tag else 'No Product Name Available'
        
        # Extract the image URL
        img_tag = soup.find('img', class_='gallery-placeholder__image') 
        img_url = img_tag['src'] if img_tag and 'src' in img_tag.attrs else 'No Image Available'
        
        # Extract the category
        # category_tag = soup.find('span', class_='category')
        # category = category_tag.text.strip() if category_tag else 'No Category Available'
        category= 'Men - Bags'

        # Extract the subcategory
        # subcategory_tag = soup.find('span', class_='subcategory')
        # subcategory = subcategory_tag.text.strip() if subcategory_tag else 'No Subcategory Available'
        subcategory= ''

        # Extract the color
        product_color_tag = soup.find('div', class_='value', itemprop='coach_style_color') 
        product_color = product_color_tag.text.strip() if product_color_tag else 'No Color Available'
        
        # Extract the current price
        current_price_tag = soup.find('span', class_='price-wrapper', attrs={'data-price-type': "finalPrice"})
        current_price = current_price_tag.text.strip() if current_price_tag else 'No Current Price Available'
        
        # Extract the original price
        original_price_tag = soup.find('span', class_='price-wrapper', attrs={'data-price-type': "oldPrice"}) 
        original_price = original_price_tag.text.strip() if original_price_tag else 'No Original Price Available'
        
        # Extract the description
        description_tag = soup.find('div', class_='product attribute overview')
        description = description_tag.text.strip() if description_tag else 'No Description Available'
        
        # Extract the details
        details_tag = soup.find('div', class_='product attribute description')
        details = details_tag.text.strip() if details_tag else 'No Details Available'
        
        # Extract the tag (internal links within the product page)
        tag_tag = soup.find('a', class_='internal-link')
        tag = tag_tag['href'] if tag_tag and 'href' in tag_tag.attrs else 'No Tag Available'
        
        # Extract the status (OOS/Pre-order, etc.)
        status_tag = soup.find('div', class_='status')
        status = status_tag.text.strip() if status_tag else 'No Status Available'
        
        # Extract the remarks (Stock arrival period)
        remarks_tag = soup.find('div', class_='product-info-stock-sku')
        remarks = remarks_tag.text.strip() if remarks_tag else 'No Remarks Available'
        
        # Get the current timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create a DataFrame for the current product
        data = {
            'Link': [product_link],
            'Name': [product_name],
            'Image': [img_url],
            'Category': [category],
            'Subcategory': [subcategory],
            'Color': [product_color],
            'Current Price': [current_price],
            'Original Price': [original_price],
            'Description': [description],
            'Details': [details],
            'Tag': [tag],
            'Status': [status],
            'Remarks': [remarks],
            'Timestamp': [timestamp]
        }
        temp_mb = pd.DataFrame(data)
        
        # Concatenate the current product DataFrame to the main DataFrame
        mb = pd.concat([mb, temp_mb], ignore_index=True)
    else:
        print(f'Failed to retrieve the webpage: {url}')

# Display the DataFrame
# print(mb)
# mb

mb['Remarks']=mb['Remarks'].transform(lambda x: ' '.join(x.split()).strip() if x else x)
mb['Details']=mb['Details'].transform(lambda x: x.replace('\n', ',') if x else x)
mb 


final=pd.concat([ws, wb, ms, mb], ignore_index=True) 
final.to_excel('scraping111.xlsx', index=False) 
final
